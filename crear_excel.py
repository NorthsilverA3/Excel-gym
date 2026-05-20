import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear un nuevo libro
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Datos de configuración para cada semana
semanas_config = [
    {"nombre": "Semana 1", "repeticiones": 12, "porcentaje": 0.7, "descanso_normal": "01:15:00", "descanso_abdomen": "00:30:00"},
    {"nombre": "Semana 2", "repeticiones": 10, "porcentaje": 0.75, "descanso_normal": "01:15:00", "descanso_abdomen": "01:00:00"},
    {"nombre": "semana 3", "repeticiones": 8, "porcentaje": 0.80, "descanso_normal": "01:15:00", "descanso_abdomen": "01:00:00"},
    {"nombre": "semana 4", "repeticiones": 6, "porcentaje": 0.85, "descanso_normal": "02:30:00", "descanso_abdomen": "02:30:00"},
    {"nombre": "semana 5", "repeticiones": 15, "porcentaje": 0.60, "descanso_normal": "01:00:00", "descanso_abdomen": "01:00:00"},
]

# Estructura de ejercicios por día
ejercicios_por_dia = {
    "LUNES": [
        ("Espalda", "Remo barra/mancuerna", 3, 12, 0.7, 2, "01:15:00"),
        ("Espalda", "Jalon al pecho", 3, 12, 0.7, 2, "01:15:00"),
        ("Espalda", "Remo autraliano", 3, 12, 0.7, 2, "01:15:00"),
        ("Biceps", "Curl de biceps maquina", 2, 12, 0.7, 1, "01:15:00"),
        ("Biceps", "Curl TRX", 2, 12, 0.7, 1, "01:15:00"),
        ("Abdomen", "Plancha Lateral", 3, "30\"", None, None, "00:30:00"),
        ("Abdomen", "Crunch Abdominal", 3, 30, None, None, "01:00:00"),
    ],
    "MARTES": [
        ("Cuadriceps", "Sentadilla libre", 3, 12, 0.7, 2, "01:15:00"),
        ("Cuadriceps", "Sentadilla Hack", 3, 12, 0.7, 2, "01:15:00"),
        ("Cuadriceps", "Prensa Inclinada", 3, 12, 0.7, 2, "01:15:00"),
        ("Isquios", "Maquina flexor acostado", 3, 12, 0.7, 2, "01:15:00"),
        ("Cuadriceps", "Maquina Extensor Sentado", 3, 12, 0.7, 2, "01:15:00"),
        ("Pantorrilla", "Elevacion talones", 6, 12, 0.7, 1, "01:15:00"),
    ],
    "MIERCOLES": [
        ("Pectoral", "Press de banca", 3, 12, 0.7, 2, "01:15:00"),
        ("Pectoral", "Flexiones", 3, 12, 0.7, 2, "01:15:00"),
        ("Pectoral", "Pecdeck", 3, 12, 0.7, 2, "01:15:00"),
        ("Deltoides", "Elevaciones laterales", 4, 12, 0.7, 1, "01:15:00"),
        ("Triceps", "Extension de brazos polea", 3, 12, 0.7, 1, "01:15:00"),
        ("Triceps", "Fondos para triceps", 3, 12, 0.7, 1, "01:15:00"),
        ("Abdomen", "Abdominales Rodillo", 3, 20, None, None, "01:00:00"),
        ("Abdomen", "Elevacion de piernas banco", 3, 20, None, None, "01:00:00"),
    ],
    "JUEVES": [
        ("Isquios", "Peso muerto", 3, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Puente", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Bulgara", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Set Ups", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Patada para gluteo", 2, 12, 0.7, 2, "01:15:00"),
        ("Isquios", "Flexor acostado isquios", 3, 12, 0.7, 2, "01:15:00"),
    ],
    "VIERNES": [
        ("Deltoides", "Press militar", 2, 12, 0.7, 1, "01:15:00"),
        ("Deltoides", "Facepull", 2, 12, 0.7, 1, "01:15:00"),
        ("Espalda", "Remo australiano", 3, 12, 0.7, 2, "01:15:00"),
        ("Pectoral", "Flexiones", 3, 12, 0.7, 2, "01:15:00"),
        ("Biceps", "Curl martillo", 2, 12, 0.7, 1, "01:15:00"),
        ("Triceps", "Extension brazos", 3, 12, 0.7, 1, "01:15:00"),
        ("Abdomen", "Elevacion rodillas en paralelas", 3, 20, None, None, "01:00:00"),
        ("Abdomen", "Torsion de tronco", 3, 30, None, None, "01:00:00"),
    ],
    "SABADO": [
        ("Cuadriceps", "Sentadilla", 2, 12, 0.7, 2, "01:15:00"),
        ("Cuadriceps", "Zancada", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Bulgara", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Zumo", 2, 12, 0.7, 2, "01:15:00"),
        ("Gluteo", "Puente a una pierna", 2, 12, 0.7, 2, "01:15:00"),
        ("Isquios", "Flexor isquio mancuerna", 3, 12, 0.7, 2, "01:15:00"),
        ("Pantorrilla", "Elevacion talones (adicional)", 3, 12, 0.7, 1, "01:15:00"),
    ],
}

# Mapeo de días a números
dias_numero = {
    "LUNES": "DIA 1",
    "MARTES": "DIA 2",
    "MIERCOLES": "DIA 3",
    "JUEVES": "DIA 4",
    "VIERNES": "DIA 5",
    "SABADO": "DIA 6",
}

def crear_hoja_semana(wb, config):
    ws = wb.create_sheet(config["nombre"])
    
    # Estilos
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=12)
    
    # Encabezado
    ws['A1'] = "UNIVERSIDAD INDUSTRIAL DE SANTANDER"
    ws['A2'] = "Plan de Entrenamiento Mensual"
    ws['A3'] = "NOMBRE:"
    ws['E3'] = "EDAD:"
    ws['G3'] = "EXP GYM:"
    ws['I3'] = "Patologias:"
    ws['A4'] = "OBJETIVO:"
    ws['B4'] = "Aumento de masa muscular general y gluteo."
    ws['A5'] = "OBSERVACION:"
    ws['B5'] = "Volumen alto con percentajes de carga adaptados. RIR 1-3 (0-2 en brazos y hombros)."
    
    # Aplicar estilos al encabezado
    for row in range(1, 6):
        for col in range(1, 15):
            ws.cell(row, col).font = header_font
    
    row = 7
    dia_num = 1
    
    # Iterar por cada día
    for dia, ejercicios in ejercicios_por_dia.items():
        # Encabezado del día
        ws.cell(row, 3).value = dia
        ws.cell(row, 3).font = header_font
        row += 1
        
        # Encabezados de columnas
        ws.cell(row, 1).value = dias_numero[dia]
        ws.cell(row, 2).value = "GM"
        ws.cell(row, 3).value = "Ejercicio"
        ws.cell(row, 4).value = "Series"
        ws.cell(row, 5).value = "Reales"
        ws.cell(row, 6).value = "Repeticiones"
        ws.cell(row, 7).value = "Reales"
        ws.cell(row, 8).value = "Peso"
        ws.cell(row, 9).value = "Porcentaje"
        ws.cell(row, 10).value = "RIR"
        ws.cell(row, 11).value = "Descanso"
        
        for col in range(1, 12):
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).border = border
        
        row += 1
        
        # Agregar ejercicios
        for gm, ejercicio, series, reps, porcentaje, rir, descanso in ejercicios:
            ws.cell(row, 2).value = gm
            ws.cell(row, 3).value = ejercicio
            ws.cell(row, 4).value = series
            
            # Ajustar valores según la semana
            if isinstance(reps, str):  # Si es tiempo
                ws.cell(row, 6).value = reps
            else:
                ws.cell(row, 6).value = config["repeticiones"] if reps == 12 else reps
            
            if porcentaje is not None:
                ws.cell(row, 9).value = config["porcentaje"]
            
            if rir is not None:
                ws.cell(row, 10).value = rir
            
            # Determinar descanso según el ejercicio
            if "Abdomen" in gm:
                ws.cell(row, 11).value = config["descanso_abdomen"]
            else:
                ws.cell(row, 11).value = config["descanso_normal"]
            
            if "Corporal" in str(descanso) or gm == "Abdomen":
                ws.cell(row, 8).value = "Corporal"
            
            for col in range(1, 12):
                ws.cell(row, col).border = border
            
            row += 1
        
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['K'].width = 15
    
    return ws

# Crear hojas para cada semana
for config in semanas_config:
    crear_hoja_semana(wb, config)

# Crear hoja de EJERCICIOS
ws_ejercicios = wb.create_sheet("EJERCICIOS")
ws_ejercicios['A1'] = "EJERCICIOS - REFERENCIA"
ws_ejercicios['A1'].font = Font(bold=True, size=12)

# Guardar el archivo
wb.save("Programacion de rutina_modificada.xlsx")
print("Archivo creado exitosamente: Programacion de rutina_modificada.xlsx")
