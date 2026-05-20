import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# Crear workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Configuración de semanas
semanas = [
    {"nombre": "Semana 1", "reps": 12, "porcentaje": 0.70, "descanso": "01:15:00"},
    {"nombre": "Semana 2", "reps": 10, "porcentaje": 0.75, "descanso": "01:15:00"},
    {"nombre": "Semana 3", "reps": 8, "porcentaje": 0.80, "descanso": "01:15:00"},
    {"nombre": "Semana 4", "reps": 6, "porcentaje": 0.85, "descanso": "02:30:00"},
    {"nombre": "Semana 5", "reps": 15, "porcentaje": 0.60, "descanso": "01:00:00"},
]

# Pesos por ejercicio (1RM)
pesos_1rm = {
    "Press de banca": 80,
    "Jalon al pecho": 70,
    "Sentadilla libre": 100,
    "Prensa": 120,
    "Hack": 110,
}

# Estructura de entrenamientos (5 días)
entrenamientos = {
    "LUNES": {
        "principal": "Pecho",
        "secundario": "Deltoides",
        "ejercicios": [
            ("Pecho", "Press de banca", 4, 80),
            ("Pecho", "Flexiones", 4, None),
            ("Pecho", "Pecdeck", 4, None),
            ("Pecho", "Aperturas con mancuernas", 4, None),
            ("Deltoides", "Elevaciones laterales", 2, None),
            ("Deltoides", "Press militar", 2, None),
        ]
    },
    "MARTES": {
        "principal": "Espalda",
        "secundario": "Bíceps",
        "ejercicios": [
            ("Espalda", "Jalon al pecho", 4, 70),
            ("Espalda", "Remo barra", 4, None),
            ("Espalda", "Remo australiano", 4, None),
            ("Espalda", "Pullover", 4, None),
            ("Bíceps", "Curl de bíceps", 2, None),
            ("Bíceps", "Curl martillo", 2, None),
        ]
    },
    "MIÉRCOLES": {
        "principal": "Piernas",
        "secundario": "Isquios",
        "ejercicios": [
            ("Cuádriceps", "Sentadilla libre", 4, 100),
            ("Cuádriceps", "Prensa", 4, 120),
            ("Cuádriceps", "Hack", 4, 110),
            ("Cuádriceps", "Extensor sentado", 4, None),
            ("Isquios", "Peso muerto", 2, None),
            ("Isquios", "Flexor acostado", 2, None),
        ]
    },
    "JUEVES": {
        "principal": "Hombro",
        "secundario": "Tríceps",
        "ejercicios": [
            ("Deltoides", "Press militar", 4, None),
            ("Deltoides", "Elevaciones laterales", 4, None),
            ("Deltoides", "Facepull", 4, None),
            ("Deltoides", "Aperturas en máquina", 4, None),
            ("Tríceps", "Extension polea", 2, None),
            ("Tríceps", "Fondos", 2, None),
        ]
    },
    "VIERNES": {
        "principal": "Glúteo",
        "secundario": "Abdomen",
        "ejercicios": [
            ("Glúteo", "Puente", 4, None),
            ("Glúteo", "Sentadilla búlgara", 4, None),
            ("Glúteo", "Patada de glúteo", 4, None),
            ("Glúteo", "Hip thrust", 4, None),
            ("Abdomen", "Abdominales rodillo", 2, None),
            ("Abdomen", "Crunch", 2, None),
        ]
    }
}

dias_semana = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]

# Crear estilos
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
center_align = Alignment(horizontal='center', vertical='center')

def crear_hoja_semana(wb, semana_config):
    ws = wb.create_sheet(semana_config["nombre"])
    
    # Encabezados principales
    ws['A1'] = "UNIVERSIDAD INDUSTRIAL DE SANTANDER"
    ws['A1'].font = title_font
    
    ws['A2'] = "Plan de Entrenamiento Mensual"
    ws['A2'].font = Font(bold=True, size=12)
    
    ws['A3'] = "NOMBRE:"
    ws['E3'] = "EDAD:"
    ws['G3'] = "EXP GYM:"
    ws['I3'] = "Patologías:"
    
    ws['A4'] = "OBJETIVO:"
    ws['B4'] = "Aumento de masa muscular general y glúteo"
    
    ws['A5'] = "OBSERVACIÓN:"
    ws['B5'] = "Volumen alto con porcentajes de carga adaptados. RIR 1-3"
    
    fila = 7
    
    for dia in dias_semana:
        # Encabezado del día
        ws[f'C{fila}'] = dia
        ws[f'C{fila}'].font = header_font
        ws[f'C{fila}'].alignment = center_align
        fila += 1
        
        # Headers de columnas
        ws[f'A{fila}'] = "DÍA"
        ws[f'B{fila}'] = "GRUPO"
        ws[f'C{fila}'] = "EJERCICIO"
        ws[f'D{fila}'] = "SERIES"
        ws[f'E{fila}'] = "REPS"
        ws[f'F{fila}'] = "PESO (kg)"
        ws[f'G{fila}'] = "%"
        ws[f'H{fila}'] = "RIR"
        ws[f'I{fila}'] = "DESCANSO"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{fila}'].font = header_font
            ws[f'{col}{fila}'].border = border
            ws[f'{col}{fila}'].alignment = center_align
        
        fila += 1
        
        # Ejercicios del día
        ejercicios = entrenamientos[dia]["ejercicios"]
        for idx, (grupo, ejercicio, series, peso_1rm) in enumerate(ejercicios):
            ws[f'A{fila}'] = entrenamientos[dia]["principal"] if idx < 4 else entrenamientos[dia]["secundario"]
            ws[f'B{fila}'] = grupo
            ws[f'C{fila}'] = ejercicio
            ws[f'D{fila}'] = series
            ws[f'E{fila}'] = semana_config["reps"]
            
            if peso_1rm:
                peso_calculado = peso_1rm * semana_config["porcentaje"]
                ws[f'F{fila}'] = round(peso_calculado, 1)
            
            ws[f'G{fila}'] = semana_config["porcentaje"]
            
            if idx < 4:  # Principal
                ws[f'H{fila}'] = 2
            else:  # Secundario
                ws[f'H{fila}'] = 1
            
            ws[f'I{fila}'] = semana_config["descanso"]
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{fila}'].border = border
                ws[f'{col}{fila}'].alignment = center_align
            
            fila += 1
        
        fila += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    
    return ws

# Crear hojas para cada semana
for semana_config in semanas:
    crear_hoja_semana(wb, semana_config)

# Crear hoja de ejercicios
ws_ejercicios = wb.create_sheet("EJERCICIOS")
ws_ejercicios['A1'] = "REFERENCIA DE EJERCICIOS"
ws_ejercicios['A1'].font = title_font

# Guardar archivo
ruta = "Programacion de rutina_modificada.xlsx"
wb.save(ruta)
print(f"Archivo creado: {ruta}")
